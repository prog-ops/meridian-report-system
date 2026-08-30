import csv
import os
import re


def norm_str(s):
    if s is None:
        return ""
    s = str(s).strip()
    s = s.replace("–", "-").replace("—", "-")
    s = re.sub(r"\s+", " ", s)
    return s


def parse_number(s):
    if s is None:
        return None
    t = str(s)
    t = t.replace("SGD", "").replace("bps", "").replace("%", "").replace("yrs", "")
    t = t.replace(",", "")
    m = re.search(r"-?\d+(?:\.\d+)?", t)
    if not m:
        return None
    return float(m.group())


def read_csv_expected(path):
    rows = []
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(
                {
                    "Section": row.get("Section", "").strip(),
                    "Metric": row.get("Metric", "").strip(),
                    "Value": row.get("Value", "").strip(),
                    "Limit": row.get("Limit", "").strip(),
                    "Utilization": row.get("Utilization", "").strip(),
                    "Status": row.get("Status", "").strip(),
                }
            )
    return rows


def read_xlsx_expected(path):
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    rows = []
    header_found = False

    for row in ws.iter_rows(values_only=True):
        vals = ["" if c is None else str(c).strip() for c in row]

        if not header_found:
            if len(vals) >= 2 and vals[0].lower() == "section" and vals[1].lower() == "metric":
                header_found = True
            continue

        if len(vals) >= 6 and vals[1]:
            rows.append(
                {
                    "Section": vals[0],
                    "Metric": vals[1],
                    "Value": vals[2],
                    "Limit": vals[3],
                    "Utilization": vals[4],
                    "Status": vals[5],
                }
            )

    return rows


def load_expected(path, fallback_path):
    if path and os.path.exists(path):
        if path.lower().endswith((".xlsx", ".xlsm")):
            return read_xlsx_expected(path)
        return read_csv_expected(path)

    if fallback_path and os.path.exists(fallback_path):
        return read_csv_expected(fallback_path)

    raise FileNotFoundError("No answer key found")


def reconcile_figures(figures, expected_rows):
    actual_map = {}
    for f in figures:
        key = norm_str(f["section"]) + "|" + norm_str(f["metric"])
        actual_map[key] = f

    items = []
    overall_pass = True

    fields = ["Value", "Limit", "Utilization", "Status"]

    for exp in expected_rows:
        key = norm_str(exp["Section"]) + "|" + norm_str(exp["Metric"])
        act = actual_map.get(key)

        if not act:
            overall_pass = False
            items.append(
                {
                    "section": exp["Section"],
                    "metric": exp["Metric"],
                    "found": False,
                    "pass": False,
                }
            )
            continue

        item = {
            "section": exp["Section"],
            "metric": exp["Metric"],
            "found": True,
            "fields": {},
        }

        row_pass = True
        for field in fields:
            exp_v = exp.get(field, "")
            act_v = act.get(field.lower(), "")
            passed = norm_str(exp_v) == norm_str(act_v)

            delta = None
            if field in {"Value", "Utilization"}:
                exp_n = parse_number(exp_v)
                act_n = parse_number(act_v)
                if exp_n is not None and act_n is not None:
                    delta = act_n - exp_n

            item["fields"][field.lower()] = {
                "expected": exp_v,
                "actual": act_v,
                "pass": passed,
                "delta": delta,
            }
            row_pass = row_pass and passed

        item["pass"] = row_pass
        overall_pass = overall_pass and row_pass
        items.append(item)

    return {
        "overall_pass": overall_pass,
        "items": items,
    }


def traceability_check(figures, G):
    issues = []

    for f in figures:
        if not f.get("graph_path"):
            issues.append({"figure": f.get("figure"), "issue": "missing graph_path"})

        citations = f.get("citations", [])
        if not citations:
            issues.append({"figure": f.get("figure"), "issue": "missing citations"})

        for c in citations:
            if not c:
                continue
            src = c.get("source_doc") or ""
            chunk = c.get("chunk_id")
            if src.endswith("sample_fund_guidelines.pdf") and chunk:
                if f"Chunk:{chunk}" not in G.nodes:
                    issues.append(
                        {
                            "figure": f.get("figure"),
                            "issue": f"guideline chunk not in graph: {chunk}",
                        }
                    )

    return {
        "passed": len(issues) == 0,
        "figures_checked": len(figures),
        "issues": issues,
    }